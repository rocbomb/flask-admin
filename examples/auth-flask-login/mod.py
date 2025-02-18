from openpyxl import Workbook


import openpyxl
import json



# 初始化一个空列表，用来存储学生信息
def readXlsx(filename):
    students_data = []

    # 打开 Excel 文件
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    # 遍历每一行数据，跳过第一行（标题行）
    for row in sheet.iter_rows(values_only=True, min_row=2):
        if row[1] == None:
            continue
        # 将学生信息存储为字典
        student_data = {
            "id": int(row[0]),
            "name": row[1],
        }
        # 将学生信息字典添加到列表中
        students_data.append(student_data)
    # 关闭 Excel 文件
    wb.close()
    return students_data
s0 = readXlsx('xlsx/student.xlsx')
s1 = readXlsx('xlsx/001.xlsx')

for s in s1:
    for _s in s0:
        if _s['name'] == s['name']:
            s['id'] = _s['id']

# 创建一个新的工作簿
wb = Workbook()

# 选择活动工作表
ws = wb.active

# 给工作表命名
ws.title = "Sheet1"

# 写入数据
ws['A1'] = 'id'
ws['B1'] = 'name'

# 添加数据
data = []
for s in s1:
    data.append([s['id'], s['name']])  # 每行数据是一个列表，包含多个单元格的值

for row in data:
    ws.append(row)  # 逐行追加数据

# 保存工作簿
wb.save('output.xlsx')